from fastapi import FastAPI, APIRouter, HTTPException
from fastapi.responses import FileResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime
import PyPDF2
import asyncio
import aiofiles
import openpyxl
import xlrd
from docx import Document
import docx2txt
from striprtf.striprtf import rtf_to_text

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Create the main app without a prefix
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# PDF base path
DOCUMENT_BASE_PATH = Path("/var/www/html/pdf")

# Supported file extensions
SUPPORTED_EXTENSIONS = {'.pdf', '.xlsx', '.xls', '.docx', '.doc', '.rtf', '.txt'}

# Models
class FileItem(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    path: str
    type: str  # 'file' or 'folder'
    size: Optional[int] = None
    modified: Optional[datetime] = None
    parent_path: str

class SearchResult(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    file_path: str
    file_name: str
    content_match: Optional[str] = None
    match_type: str  # 'filename' or 'content'

class IndexedFile(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    file_path: str
    file_name: str
    content: str
    indexed_at: datetime = Field(default_factory=lambda: datetime.now())

# Utility functions
def get_file_info(file_path: Path) -> Dict[str, Any]:
    """Get file information"""
    try:
        stat = file_path.stat()
        return {
            "name": file_path.name,
            "path": str(file_path),
            "type": "folder" if file_path.is_dir() else "file",
            "size": stat.st_size if file_path.is_file() else None,
            "modified": datetime.fromtimestamp(stat.st_mtime),
            "parent_path": str(file_path.parent)
        }
    except Exception as e:
        logger.error(f"Error getting file info for {file_path}: {e}")
        return None

async def extract_pdf_text(file_path: Path) -> str:
    """Extract text content from PDF"""
    try:
        text = ""
        async with aiofiles.open(file_path, 'rb') as file:
            content = await file.read()
            
        # Use PyPDF2 to extract text
        import io
        pdf_reader = PyPDF2.PdfReader(io.BytesIO(content))
        
        for page in pdf_reader.pages:
            text += page.extract_text() + "\n"
            
        return text.strip()
    except Exception as e:
        logger.error(f"Error extracting text from PDF {file_path}: {e}")
        return ""

async def extract_excel_text(file_path: Path) -> str:
    """Extract text content from Excel files"""
    try:
        text = ""
        if file_path.suffix.lower() == '.xlsx':
            wb = openpyxl.load_workbook(file_path, data_only=True)
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                text += f"Sheet: {sheet_name}\n"
                for row in sheet.iter_rows(values_only=True):
                    row_text = " | ".join([str(cell) if cell is not None else "" for cell in row])
                    if row_text.strip():
                        text += row_text + "\n"
                text += "\n"
        elif file_path.suffix.lower() == '.xls':
            workbook = xlrd.open_workbook(file_path)
            for sheet_idx in range(workbook.nsheets):
                sheet = workbook.sheet_by_index(sheet_idx)
                text += f"Sheet: {sheet.name}\n"
                for row_idx in range(sheet.nrows):
                    row_text = " | ".join([str(sheet.cell_value(row_idx, col_idx)) for col_idx in range(sheet.ncols)])
                    if row_text.strip():
                        text += row_text + "\n"
                text += "\n"
        return text.strip()
    except Exception as e:
        logger.error(f"Error extracting text from Excel {file_path}: {e}")
        return ""

async def extract_word_text(file_path: Path) -> str:
    """Extract text content from Word documents"""
    try:
        if file_path.suffix.lower() == '.docx':
            doc = Document(file_path)
            text = ""
            for paragraph in doc.paragraphs:
                text += paragraph.text + "\n"
            return text.strip()
        elif file_path.suffix.lower() == '.doc':
            # Use docx2txt for .doc files
            text = docx2txt.process(str(file_path))
            return text.strip() if text else ""
        return ""
    except Exception as e:
        logger.error(f"Error extracting text from Word document {file_path}: {e}")
        return ""

async def extract_rtf_text(file_path: Path) -> str:
    """Extract text content from RTF files"""
    try:
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as file:
            rtf_content = file.read()
        text = rtf_to_text(rtf_content)
        return text.strip()
    except Exception as e:
        logger.error(f"Error extracting text from RTF {file_path}: {e}")
        return ""

async def extract_text_file(file_path: Path) -> str:
    """Extract text content from plain text files"""
    try:
        async with aiofiles.open(file_path, 'r', encoding='utf-8', errors='ignore') as file:
            content = await file.read()
        return content.strip()
    except Exception as e:
        logger.error(f"Error reading text file {file_path}: {e}")
        return ""

async def extract_document_text(file_path: Path) -> str:
    """Extract text from any supported document format"""
    extension = file_path.suffix.lower()
    
    if extension == '.pdf':
        return await extract_pdf_text(file_path)
    elif extension in ['.xlsx', '.xls']:
        return await extract_excel_text(file_path)
    elif extension in ['.docx', '.doc']:
        return await extract_word_text(file_path)
    elif extension == '.rtf':
        return await extract_rtf_text(file_path)
    elif extension == '.txt':
        return await extract_text_file(file_path)
    else:
        logger.warning(f"Unsupported file format: {extension}")
        return ""

# API Routes
@api_router.get("/")
async def root():
    return {"message": "PDF Search System API"}

@api_router.get("/files/tree")
async def get_file_tree(path: str = ""):
    """Get folder structure and files"""
    try:
        current_path = DOCUMENT_BASE_PATH / path if path else DOCUMENT_BASE_PATH
        
        if not current_path.exists():
            raise HTTPException(status_code=404, detail="Path not found")
        
        items = []
        
        # Get all items in current directory
        try:
            for item_path in sorted(current_path.iterdir()):
                if item_path.name.startswith('.'):
                    continue
                    
                file_info = get_file_info(item_path)
                if file_info:
                    # For relative path calculation
                    relative_path = str(item_path.relative_to(DOCUMENT_BASE_PATH))
                    file_info["path"] = relative_path
                    file_info["parent_path"] = str(item_path.parent.relative_to(DOCUMENT_BASE_PATH)) if item_path.parent != DOCUMENT_BASE_PATH else ""
                    
                    items.append(FileItem(**file_info))
                    
        except PermissionError:
            raise HTTPException(status_code=403, detail="Permission denied")
            
        return {"items": items, "current_path": path}
        
    except Exception as e:
        logger.error(f"Error getting file tree: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/files/serve/{file_path:path}")
async def serve_document(file_path: str):
    """Serve document file for preview/download"""
    try:
        full_path = DOCUMENT_BASE_PATH / file_path
        
        if not full_path.exists() or not full_path.is_file():
            raise HTTPException(status_code=404, detail="File not found")
            
        extension = full_path.suffix.lower()
        if extension not in SUPPORTED_EXTENSIONS:
            raise HTTPException(status_code=400, detail=f"File type {extension} not supported")
        
        # Set appropriate media type based on file extension
        media_types = {
            '.pdf': 'application/pdf',
            '.xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            '.xls': 'application/vnd.ms-excel',
            '.docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
            '.doc': 'application/msword',
            '.rtf': 'application/rtf',
            '.txt': 'text/plain'
        }
        
        media_type = media_types.get(extension, 'application/octet-stream')
        
        return FileResponse(
            path=str(full_path),
            media_type=media_type,
            filename=full_path.name
        )
        
    except Exception as e:
        logger.error(f"Error serving file {file_path}: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/files/index")
async def index_documents():
    """Index all supported document files for content search"""
    try:
        indexed_count = 0
        
        # Clear existing index
        await db.indexed_files.delete_many({})
        
        # Walk through all supported document files
        for extension in SUPPORTED_EXTENSIONS:
            for doc_path in DOCUMENT_BASE_PATH.rglob(f"*{extension}"):
                try:
                    # Extract text content
                    content = await extract_document_text(doc_path)
                    
                    if content:
                        # Store in database
                        indexed_file = IndexedFile(
                            file_path=str(doc_path.relative_to(DOCUMENT_BASE_PATH)),
                            file_name=doc_path.name,
                            content=content
                        )
                        
                        await db.indexed_files.insert_one(indexed_file.dict())
                        indexed_count += 1
                        
                        logger.info(f"Indexed: {doc_path.name}")
                        
                except Exception as e:
                    logger.error(f"Error indexing {doc_path}: {e}")
                    continue
                    
        return {"message": f"Indexed {indexed_count} document files"}
        
    except Exception as e:
        logger.error(f"Error during indexing: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/search")
async def search_files(q: str, limit: int = 50):
    """Search files by name and content"""
    try:
        if not q or len(q.strip()) < 2:
            return {"results": []}
            
        search_term = q.strip().lower()
        results = []
        
        # Search by filename
        for pdf_path in DOCUMENT_BASE_PATH.rglob("*.pdf"):
            if search_term in pdf_path.name.lower():
                results.append(SearchResult(
                    file_path=str(pdf_path.relative_to(DOCUMENT_BASE_PATH)),
                    file_name=pdf_path.name,
                    match_type="filename"
                ))
        
        # Search by content in indexed files
        content_matches = await db.indexed_files.find({
            "content": {"$regex": search_term, "$options": "i"}
        }).to_list(limit)
        
        for match in content_matches:
            # Extract snippet around match
            content = match.get("content", "")
            lower_content = content.lower()
            match_index = lower_content.find(search_term)
            
            if match_index >= 0:
                start = max(0, match_index - 100)
                end = min(len(content), match_index + 100)
                snippet = content[start:end]
                
                # Check if already in results (from filename search)
                existing = next((r for r in results if r.file_path == match["file_path"]), None)
                if existing:
                    existing.content_match = f"...{snippet}..."
                    existing.match_type = "both"
                else:
                    results.append(SearchResult(
                        file_path=match["file_path"],
                        file_name=match["file_name"],
                        content_match=f"...{snippet}...",
                        match_type="content"
                    ))
        
        # Limit results
        results = results[:limit]
        
        return {"results": results, "total": len(results)}
        
    except Exception as e:
        logger.error(f"Error searching: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()